#!/usr/bin/env python3
"""
Usage:
  python process-brand-labels.py <brands.csv> <products.csv>

Reads two semicolon-delimited CSVs:
  - brands.csv   : code;label-{locale};...
  - products.csv : code;brand;product_name-{locale}-{label};short_product_name-{locale}-{label};...

Outputs two files in the same folder as products.csv:
  - <products>_processed.xlsx         : processed products only
  - <products>_combined_processed.xlsx: processed products (sheet 1) + brands (sheet 2)
"""

import csv
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

DUPLICATE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def parse_csv(path: str) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        rows = [dict(row) for row in reader]
        headers = [h for h in (reader.fieldnames or []) if h]
    return headers, rows


def build_brand_label_map(brand_rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {
        row["code"].strip(): {
            key.replace("label-", ""): value
            for key, value in row.items()
            if key.startswith("label-") and value.strip()
        }
        for row in brand_rows
        if row.get("code", "").strip()
    }


def extract_locale(col_name: str) -> str | None:
    # 'short_product_name-nl_NL-beltegoed_nl' -> 'nl_NL'
    parts = col_name.split("-")
    return parts[1] if len(parts) >= 2 else None


def find_duplicate_indices(product_rows: list[dict[str, str]]) -> set[int]:
    brand_groups: dict[str, list[int]] = defaultdict(list)
    for i, row in enumerate(product_rows):
        brand = row.get("brand", "")
        if brand:
            brand_groups[brand].append(i)

    duplicate_indices: set[int] = set()

    for indices in brand_groups.values():
        if len(indices) <= 1:
            continue

        parentless = [i for i in indices if not product_rows[i].get("parent", "")]
        children = [i for i in indices if product_rows[i].get("parent", "")]

        # Multiple parentless products with same brand → all are duplicates
        if len(parentless) > 1:
            duplicate_indices.update(parentless)

        # Children pointing to different parents → all children are duplicates
        # (the single parentless main product is intentionally left out)
        child_parents = {product_rows[i].get("parent", "") for i in children}
        if len(child_parents) > 1:
            duplicate_indices.update(children)

    return duplicate_indices


def write_products_sheet(ws, product_headers: list[str], product_rows: list[dict[str, str]], brand_label_map: dict) -> tuple[int, int]:
    short_name_cols = [h for h in product_headers if h.startswith("short_product_name-")]
    duplicate_indices = find_duplicate_indices(product_rows)
    ws.append(product_headers)

    duplicate_count = 0
    filled_count = 0

    for i, row in enumerate(product_rows):
        updated = dict(row)

        if i in duplicate_indices:
            for col in short_name_cols:
                updated[col] = ""
            duplicate_count += 1
        else:
            brand_labels = brand_label_map.get(row.get("brand", ""), {})
            for col in short_name_cols:
                locale = extract_locale(col)
                label = brand_labels.get(locale, "") if locale else ""
                if label:
                    words = label.split(" ")
                    updated[col] = " ".join(w.capitalize() for w in words) if not any(w[0].isupper() for w in words if w) else label
            filled_count += 1

        ws.append([updated.get(h, "") for h in product_headers])

        if i in duplicate_indices:
            for cell in ws[ws.max_row]:
                cell.fill = DUPLICATE_FILL

    return duplicate_count, filled_count


def write_brands_sheet(ws, brand_headers: list[str], brand_rows: list[dict[str, str]]) -> None:
    ws.append(brand_headers)
    for row in brand_rows:
        ws.append([row.get(h, "") for h in brand_headers])


def process(brands_path: str, products_path: str, output_dir: Path, stem: str) -> None:
    brand_headers, brand_rows = parse_csv(brands_path)
    product_headers, product_rows = parse_csv(products_path)

    brand_label_map = build_brand_label_map(brand_rows)

    wb = Workbook()
    wb.active.title = "Products"
    duplicate_count, filled_count = write_products_sheet(wb.active, product_headers, product_rows, brand_label_map)
    brands_ws = wb.create_sheet(title="Brands")
    write_brands_sheet(brands_ws, brand_headers, brand_rows)
    path = output_dir / f"{stem} Processed.xlsx"
    wb.save(path)

    print(f"Duplicates highlighted : {duplicate_count}")
    print(f"Short names filled     : {filled_count}")
    print(f"Output                 : {path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python process-brand-labels.py <brands.csv> <products.csv>")
        sys.exit(1)
    products_path = Path(sys.argv[2])
    stem = products_path.stem.replace(" Clean", "").replace("_Clean", "").strip()
    process(sys.argv[1], sys.argv[2], products_path.parent, stem)
